import requests
import lxml.html
import openpyxl
from openpyxl import Workbook


def parse(url):
    api = requests.get(url)
    tree = lxml.html.document_fromstring(api.text)
    words = tree.xpath('/html/body/div[4]/div[1]/div[1]/div[1]/div[2]/ul/li/a/text()')
    return words


def main():
    wb = Workbook()
    # удаляем созданный автоматически лист
    wb.remove(wb['Sheet'])
    # ws = wb.active
    for i in range(12):
        if i >= 1:
            words = parse(f'https://www.allscrabblewords.com/{i + 1}-letter-words/')

            # to excel
            # создаем лист с подходящим названием
            wb.create_sheet(f'{i+1}_words')
            # выбираем этот лист
            sheet = wb[f'{i+1}_words']
            # делаем еще один цикл по данным и данные пишем в файл
            for word in words:
                cell = sheet.cell(row=words.index(word)+1, column=1)
                cell.value = word

            # to txt
            with open(f'files/{i+1}_words.txt', 'w', encoding='utf-8') as file:
                for items in words:
                    file.write(f'{items}\n')
    wb.save('words.xlsx')


if __name__ == '__main__':
    main()
